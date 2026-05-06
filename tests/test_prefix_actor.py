"""
Unit tests for prefix_actor.py – header normalisation, prefix logic, output path.
Run with:  pytest tests/
"""
import os
import sys

import pandas as pd
import pytest

# Ensure project root is on sys.path when running from repo root
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from prefix_actor import PrefixActorError, normalize_col, process_prefix_actor


# ---------------------------------------------------------------------------
# normalize_col
# ---------------------------------------------------------------------------

class TestNormalizeCol:
    def test_lowercase(self):
        assert normalize_col("Actor") == "actor"
        assert normalize_col("TEXT") == "text"
        assert normalize_col("ACTOR") == "actor"

    def test_strips_edges(self):
        assert normalize_col("  actor  ") == "actor"
        assert normalize_col(" Text ") == "text"

    def test_collapses_internal_whitespace(self):
        assert normalize_col("ac  tor") == "ac tor"
        assert normalize_col("  Text  ") == "text"

    def test_mixed(self):
        assert normalize_col("  ACTOR  ") == "actor"
        assert normalize_col(" Text ") == "text"


# ---------------------------------------------------------------------------
# process_prefix_actor – CSV
# ---------------------------------------------------------------------------

class TestProcessPrefixActorCSV:
    def _make_csv(self, path, rows, actor_header="actor", text_header="Text"):
        df = pd.DataFrame(rows, columns=[actor_header, text_header])
        df.to_csv(path, index=False, encoding="utf-8-sig")

    def test_basic_prefix(self, tmp_path):
        csv_file = str(tmp_path / "data.csv")
        self._make_csv(csv_file, [("Alice", "Hello world")])
        out_dir = str(tmp_path / "output")

        out_path = process_prefix_actor(csv_file, out_dir)

        df_out = pd.read_csv(out_path, encoding="utf-8-sig", dtype=str, keep_default_na=False)
        assert df_out["Text"].iloc[0] == "[Alice] Hello world"

    def test_empty_actor_still_prefixes(self, tmp_path):
        csv_file = str(tmp_path / "data.csv")
        self._make_csv(csv_file, [("", "Some text")])
        out_dir = str(tmp_path / "output")

        out_path = process_prefix_actor(csv_file, out_dir)

        df_out = pd.read_csv(out_path, encoding="utf-8-sig", dtype=str, keep_default_na=False)
        assert df_out["Text"].iloc[0] == "[] Some text"

    def test_empty_text_no_trailing_space(self, tmp_path):
        csv_file = str(tmp_path / "data.csv")
        self._make_csv(csv_file, [("Alice", "")])
        out_dir = str(tmp_path / "output")

        out_path = process_prefix_actor(csv_file, out_dir)

        df_out = pd.read_csv(out_path, encoding="utf-8-sig", dtype=str, keep_default_na=False)
        assert df_out["Text"].iloc[0] == "[Alice]"

    def test_case_insensitive_header_actor(self, tmp_path):
        csv_file = str(tmp_path / "data.csv")
        self._make_csv(csv_file, [("Bob", "Hi")], actor_header="ACTOR")
        out_dir = str(tmp_path / "output")

        out_path = process_prefix_actor(csv_file, out_dir)

        df_out = pd.read_csv(out_path, encoding="utf-8-sig", dtype=str, keep_default_na=False)
        assert df_out["ACTOR"].iloc[0] == "Bob"  # actor col unchanged
        assert df_out["Text"].iloc[0] == "[Bob] Hi"

    def test_case_insensitive_header_text(self, tmp_path):
        csv_file = str(tmp_path / "data.csv")
        self._make_csv(csv_file, [("Bob", "Hi")], text_header="TEXT")
        out_dir = str(tmp_path / "output")

        out_path = process_prefix_actor(csv_file, out_dir)

        df_out = pd.read_csv(out_path, encoding="utf-8-sig", dtype=str, keep_default_na=False)
        assert df_out["TEXT"].iloc[0] == "[Bob] Hi"

    def test_header_with_extra_spaces(self, tmp_path):
        csv_file = str(tmp_path / "data.csv")
        self._make_csv(csv_file, [("Carol", "Bye")], actor_header=" actor ", text_header=" Text ")
        out_dir = str(tmp_path / "output")

        out_path = process_prefix_actor(csv_file, out_dir)

        df_out = pd.read_csv(out_path, encoding="utf-8-sig", dtype=str, keep_default_na=False)
        assert df_out[" Text "].iloc[0] == "[Carol] Bye"

    def test_output_path_is_in_output_dir(self, tmp_path):
        csv_file = str(tmp_path / "report.csv")
        self._make_csv(csv_file, [("X", "Y")])
        out_dir = str(tmp_path / "output")

        out_path = process_prefix_actor(csv_file, out_dir)

        assert os.path.dirname(os.path.abspath(out_path)) == os.path.abspath(out_dir)
        assert os.path.basename(out_path) == "report.csv"

    def test_does_not_overwrite_input(self, tmp_path):
        csv_file = str(tmp_path / "data.csv")
        self._make_csv(csv_file, [("A", "original")])
        out_dir = str(tmp_path / "output")

        process_prefix_actor(csv_file, out_dir)

        df_in = pd.read_csv(csv_file, encoding="utf-8-sig", dtype=str, keep_default_na=False)
        assert df_in["Text"].iloc[0] == "original"

    def test_missing_actor_column_raises(self, tmp_path):
        csv_file = str(tmp_path / "data.csv")
        df = pd.DataFrame([{"Text": "hello", "other": "x"}])
        df.to_csv(csv_file, index=False)
        with pytest.raises(PrefixActorError, match="actor"):
            process_prefix_actor(csv_file, str(tmp_path / "output"))

    def test_missing_text_column_raises(self, tmp_path):
        csv_file = str(tmp_path / "data.csv")
        df = pd.DataFrame([{"actor": "Alice", "other": "x"}])
        df.to_csv(csv_file, index=False)
        with pytest.raises(PrefixActorError, match="Text"):
            process_prefix_actor(csv_file, str(tmp_path / "output"))

    def test_unsupported_extension_raises(self, tmp_path):
        txt_file = str(tmp_path / "data.txt")
        with open(txt_file, "w") as f:
            f.write("actor,Text\nAlice,Hello")
        with pytest.raises(PrefixActorError, match=".txt"):
            process_prefix_actor(txt_file, str(tmp_path / "output"))

    def test_creates_output_dir_if_missing(self, tmp_path):
        csv_file = str(tmp_path / "data.csv")
        self._make_csv(csv_file, [("A", "B")])
        out_dir = str(tmp_path / "nonexistent_output")
        assert not os.path.exists(out_dir)

        process_prefix_actor(csv_file, out_dir)

        assert os.path.isdir(out_dir)

    def test_multiple_rows(self, tmp_path):
        csv_file = str(tmp_path / "data.csv")
        self._make_csv(csv_file, [("Alice", "Hi"), ("Bob", "Bye"), ("", "Empty actor")])
        out_dir = str(tmp_path / "output")

        out_path = process_prefix_actor(csv_file, out_dir)

        df_out = pd.read_csv(out_path, encoding="utf-8-sig", dtype=str, keep_default_na=False)
        assert df_out["Text"].tolist() == ["[Alice] Hi", "[Bob] Bye", "[] Empty actor"]


# ---------------------------------------------------------------------------
# process_prefix_actor – XLSX
# ---------------------------------------------------------------------------

class TestProcessPrefixActorXLSX:
    def _make_xlsx(self, path, rows, actor_header="actor", text_header="Text"):
        df = pd.DataFrame(rows, columns=[actor_header, text_header])
        df.to_excel(path, index=False)

    def test_basic_prefix_xlsx(self, tmp_path):
        xlsx_file = str(tmp_path / "data.xlsx")
        self._make_xlsx(xlsx_file, [("Alice", "Hello")])
        out_dir = str(tmp_path / "output")

        out_path = process_prefix_actor(xlsx_file, out_dir)

        assert out_path.endswith(".xlsx")
        df_out = pd.read_excel(out_path, dtype=str)
        df_out = df_out.fillna("")
        assert df_out["Text"].iloc[0] == "[Alice] Hello"

    def test_processes_first_sheet_only(self, tmp_path):
        """Only the first sheet should be processed; other sheets are untouched."""
        import openpyxl
        xlsx_file = str(tmp_path / "multi.xlsx")

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["actor", "Text"])
        ws1.append(["Alice", "Hello"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["actor", "Text"])
        ws2.append(["Bob", "World"])

        wb.save(xlsx_file)

        out_dir = str(tmp_path / "output")
        out_path = process_prefix_actor(xlsx_file, out_dir)

        import openpyxl as xl
        wb_out = xl.load_workbook(out_path)
        ws_out = wb_out.worksheets[0]
        rows = list(ws_out.values)
        # headers at rows[0], data at rows[1]
        assert rows[1][1] == "[Alice] Hello"

    def test_empty_actor_xlsx(self, tmp_path):
        xlsx_file = str(tmp_path / "data.xlsx")
        self._make_xlsx(xlsx_file, [("", "Some text")])
        out_dir = str(tmp_path / "output")

        out_path = process_prefix_actor(xlsx_file, out_dir)

        df_out = pd.read_excel(out_path, dtype=str)
        df_out = df_out.fillna("")
        assert df_out["Text"].iloc[0] == "[] Some text"

    def test_output_path_same_name_xlsx(self, tmp_path):
        xlsx_file = str(tmp_path / "report.xlsx")
        self._make_xlsx(xlsx_file, [("X", "Y")])
        out_dir = str(tmp_path / "output")

        out_path = process_prefix_actor(xlsx_file, out_dir)

        assert os.path.basename(out_path) == "report.xlsx"
